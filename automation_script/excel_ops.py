"""
Excel Operations module for reading templates and writing output files.
Supports both vertical (key-value) and multi-column (multi-site) formats.
"""
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime


# Field name mappings to normalize variations between template labels and JSON keys
FIELD_MAPPINGS = {
    # BOQ fields
    "project name": ["project_name", "project"],
    "engineering consultant": ["engineering_consultant", "consultant"],
    "name of the epc": ["epc", "epc_name", "name_of_epc"],
    "9com numbers": ["9com_numbers", "equipment_list", "applicable_9com"],
    "standards": ["standards", "applicable_standards", "list_of_standards"],
    "number of systems": ["number_of_systems", "systems_count"],
    "wattage": ["wattage", "load_value", "load", "wattage_load"],
    "load value": ["wattage", "load_value", "load"],
    "number of sites": ["number_of_sites", "sites_count", "numbers_of_sites"],
    "battery type": ["battery_type", "required_battery_type"],
    "battery autonomy": ["battery_autonomy", "autonomy", "required_battery_autonomy"],
    "battery capacity": ["battery_capacity", "capacity", "required_battery_capacity"],
    "environmental": ["environmental_conditions", "environment"],
    "temperature": ["temperature_range", "temperature"],
    "support structure": ["support_structure", "structure"],
    "specifications": ["other_specifications", "specifications", "other_equipment"],
    "other service": ["other_services", "services"],
    
    # Sizing fields - matching exact question text patterns
    "solar panels config": ["solar_panels_config", "solar_config"],
    "charge controllers config": ["charge_controllers_config", "charge_controller_config"],
    "batteries config": ["batteries_config", "battery_config"],
    "load-list": ["load_list", "loads"],
    "load list": ["load_list", "loads"],
    "future expansion": ["future_expansion_factor", "future_expansion"],
    "battery back-up": ["battery_backup_time", "backup_time"],
    "back-up time": ["battery_backup_time", "backup_time"],
    "ageing factor": ["ageing_factor", "aging_factor"],
    "design factor": ["design_factor"],
    "temperature compensation": ["temperature_compensation"],
    "other factor": ["other_battery_factors", "other_factors"],
    "computed required battery": ["computed_battery_capacity", "required_battery_capacity"],
    "required battery capacity": ["computed_battery_capacity", "required_battery_capacity"],
    "end of discharge": ["end_of_discharge_voltage", "eod_voltage"],
    "cells in series": ["cells_in_series", "number_of_cells"],
    "proposed battery cell": ["proposed_cell_capacity", "proposed_capacity"],
    "parallel sets": ["parallel_strings", "number_of_parallel"],
    "parallel strings": ["parallel_strings", "number_of_parallel"],
    "derating factors": ["derating_factors", "solar_derating"],
    "sun hours": ["sun_hours", "effective_sun_hours"],
    "future factor": ["solar_future_factor", "future_factor"],
    "formula": ["solar_sizing_formula", "sizing_formula"],
    "total daily required": ["total_daily_ah", "daily_required_ah"],
    "panels in one string": ["solar_panels_per_string", "panels_per_string"],
    "parallel solar panels": ["parallel_solar_panels", "parallel_panels"],
    "how many solar panels": ["solar_panels_per_string", "panels_per_string"],
    "how many parallel": ["parallel_solar_panels", "parallel_panels"],
    
    # SLD fields  
    "junction box": ["array_junction_boxes", "junction_boxes"],
    "charge controller type": ["charge_controller_type", "controller_type"],
    "mppt or pwm": ["charge_controller_type", "controller_type"],
    "hard-wired signals": ["hardwired_signals", "signals_to_rtu"],
    "hardwired signals": ["hardwired_signals", "signals_to_rtu"],
    "other signals": ["other_signals", "other_alarms"],
    "battery breaker box": ["battery_breaker_box", "breaker_box_required"],
    "battery breaker boxes": ["num_battery_breaker_boxes", "number_of_breaker_boxes"],
    "battery configuration": ["battery_config", "batteries_config"],
    "battery type": ["battery_type"],
    "nicd or vrla": ["battery_type"],
    "cells in series": ["cells_in_series", "number_of_cells"],
    "strings of batteries": ["battery_strings", "number_of_strings"],
    "how many strings": ["battery_strings", "number_of_strings"],
    "enclosure ip": ["battery_enclosure_rating", "enclosure_rating"],
    "nema rating": ["battery_enclosure_rating", "enclosure_rating"],
    "back-up": ["required_backup", "backup_autonomy"],
    "autonomy": ["required_backup", "backup_autonomy"],
    "panel board": ["panel_board_required", "los_required"],
    "db / los": ["panel_board_required", "los_required"],
    "los required": ["panel_board_required", "los_required"],
    "enclosure rating of los": ["los_enclosure_rating", "power_panel_rating"],
    "power panel": ["los_enclosure_rating", "power_panel_rating"],
    "breakers and ratings": ["breaker_list", "number_of_breakers"],
    "number of breakers": ["breaker_list", "number_of_breakers"],
    "notes for pv": ["pv_notes", "charge_controller_notes"],
    "notes for batteries": ["battery_notes", "enclosure_notes"],
    "other equipment": ["other_equipment", "additional_equipment"],
    "critical points": ["critical_points", "other_notes"],
}


def normalize_key(key: str) -> str:
    """Normalize a key by removing underscores and converting to lowercase."""
    return key.lower().replace("_", " ").replace("-", " ").strip()


def find_matching_data_key(template_field: str, data_keys: list) -> str | None:
    """Find a matching data key for a template field."""
    template_lower = normalize_key(template_field)
    
    # Direct match
    for key in data_keys:
        if normalize_key(key) == template_lower:
            return key
    
    # Partial match
    for key in data_keys:
        key_normalized = normalize_key(key)
        if key_normalized in template_lower or template_lower in key_normalized:
            return key
    
    # Check predefined mappings
    for template_pattern, possible_keys in FIELD_MAPPINGS.items():
        if template_pattern in template_lower:
            for possible_key in possible_keys:
                for key in data_keys:
                    if normalize_key(key) == possible_key or possible_key in normalize_key(key):
                        return key
    
    return None


def detect_template_format(ws) -> tuple[str, int]:
    """
    Detect template format and find header row.
    
    Returns:
        (format_type, header_row_index)
        format_type: 'vertical', 'multi-column', or 'horizontal'
    """
    # Check for multi-column format (row with "Item" + site names)
    for row_idx in range(1, 30):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        cell_c = ws.cell(row=row_idx, column=3).value
        
        if cell_a and cell_b:
            a_str = str(cell_a).lower()
            # Multi-column header typically has "Item" or "Question" in A
            if ("item" in a_str or "question" in a_str) and cell_b and cell_c:
                return ("multi-column", row_idx)
    
    # Check for vertical key-value format
    col_a_labels = 0
    for row in range(1, min(25, ws.max_row + 1)):
        a_val = ws.cell(row=row, column=1).value
        if a_val and isinstance(a_val, str) and len(str(a_val)) > 5 and "?" in str(a_val):
            col_a_labels += 1
    
    if col_a_labels >= 3:
        return ("vertical", 0)
    
    return ("horizontal", 0)


def fill_multicolumn_template(ws, data: dict, header_row: int) -> int:
    """
    Fill a multi-column template where each column represents a site/system.
    
    Args:
        ws: The worksheet to fill.
        data: The data dictionary with 'systems' array.
        header_row: Row index of the header row.
        
    Returns:
        Number of fields populated.
    """
    filled = 0
    
    # Get systems array
    systems = data.get("systems", [])
    if not systems:
        print("No 'systems' array found in data")
        return 0
    
    print(f"Found {len(systems)} systems to populate")
    
    # Map site names to column indices (columns B, C, D, etc. = indices 2, 3, 4)
    site_columns = {}  # {site_name: column_index}
    for col_idx in range(2, min(20, ws.max_column + 1)):
        header_val = ws.cell(row=header_row, column=col_idx).value
        if header_val:
            site_columns[str(header_val).strip()] = col_idx
    
    print(f"Template site columns: {list(site_columns.keys())}")
    
    # Match each system to a column - assign by order since template uses generic "Site 1, Site 2, Site 3"
    system_to_col = {}
    col_indices = sorted(site_columns.values())
    for i, system in enumerate(systems):
        if i < len(col_indices):
            site_name = system.get("site_name", f"System_{i+1}")
            system_to_col[site_name] = col_indices[i]
            # Also update the header with actual site name
            ws.cell(row=header_row, column=col_indices[i], value=site_name)
    
    print(f"System to column mapping: {system_to_col}")
    
    # Get data keys from first system
    if not systems:
        return 0
    data_keys = list(systems[0].keys())
    print(f"Data keys available: {data_keys}")
    
    # Fill data row by row
    for row_idx in range(header_row + 1, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if not label:
            continue
        
        # Find matching data key
        matching_key = find_matching_data_key(str(label), data_keys)
        if not matching_key:
            continue
        
        # Fill each system's value in its column
        for system in systems:
            site_name = system.get("site_name", "")
            if site_name in system_to_col:
                col_idx = system_to_col[site_name]
                value = system.get(matching_key, "")
                
                # Format complex values
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = ", ".join(f"{k}: {v}" for k, v in value.items())
                
                ws.cell(row=row_idx, column=col_idx, value=value)
                filled += 1
    
    return filled


def fill_vertical_template(ws, data: dict) -> int:
    """Fill a vertical key-value format template."""
    filled = 0
    data_keys = list(data.keys())
    
    # Handle nested 'items' or 'systems'
    if "items" in data and isinstance(data["items"], list):
        items_text = []
        for item in data["items"]:
            if isinstance(item, dict):
                item_parts = [f"{k}: {v}" for k, v in item.items() if v]
                items_text.append(" | ".join(item_parts))
            else:
                items_text.append(str(item))
        data["equipment_list"] = "\n".join(items_text)
        data_keys = list(data.keys())
    
    if "systems" in data and isinstance(data["systems"], list):
        # For vertical templates with multi-system data, show first system or summary
        if data["systems"]:
            for key, value in data["systems"][0].items():
                if key not in data:
                    data[key] = value
            data_keys = list(data.keys())
    
    # Scan rows for labels
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        
        if not cell_a or not isinstance(cell_a, str):
            continue
        
        matching_key = find_matching_data_key(cell_a, data_keys)
        
        if matching_key:
            value = data.get(matching_key, "")
            
            if isinstance(value, list):
                if all(isinstance(v, dict) for v in value):
                    formatted = []
                    for v in value:
                        formatted.append(" | ".join(f"{k}: {val}" for k, val in v.items() if val))
                    value = "\n".join(formatted)
                else:
                    value = ", ".join(str(v) for v in value)
            elif isinstance(value, dict):
                value = ", ".join(f"{k}: {v}" for k, v in value.items() if v)
            
            ws.cell(row=row, column=2, value=value)
            filled += 1
            print(f"  Mapped: '{str(cell_a)[:40]}...' <- '{matching_key}'")
    
    return filled


def fill_excel_template(template_path: str, output_path: str, data: dict | list) -> bool:
    """
    Fills the Excel template with the extracted JSON data.
    Automatically detects vertical, multi-column, or horizontal format.
    """
    try:
        if not os.path.exists(template_path):
            print(f"Template file not found: {template_path}")
            return False
        
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Normalize data
        if isinstance(data, list):
            data = {"systems": data}
        
        # Detect format
        template_format, header_row = detect_template_format(ws)
        print(f"Detected template format: {template_format}")
        
        if template_format == "multi-column" and header_row > 0:
            filled = fill_multicolumn_template(ws, data, header_row)
            print(f"Populated {filled} cells across sites")
        elif template_format == "vertical":
            filled = fill_vertical_template(ws, data)
            print(f"Populated {filled} fields")
        else:
            # Fallback to vertical
            filled = fill_vertical_template(ws, data)
            print(f"Populated {filled} fields (fallback)")
        
        wb.save(output_path)
        print(f"Successfully saved output to: {output_path}")
        return True
        
    except Exception as e:
        print(f"Error writing Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_template_columns(template_path: str) -> list[str]:
    """
    Reads the column headers from an Excel template.
    
    Args:
        template_path: Path to the Excel template.
        
    Returns:
        List of column header names.
    """
    try:
        wb = openpyxl.load_workbook(template_path, read_only=True)
        ws = wb.active
        
        headers = []
        # Check first 10 rows for headers
        for row_idx in range(1, 11):
            row_headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_headers.append(str(cell_value).strip())
            
            if len(row_headers) >= 2:  # Found a row with multiple values
                headers = row_headers
                break
        
        wb.close()
        return headers
        
    except Exception as e:
        print(f"Error reading template headers: {e}")
        return []
