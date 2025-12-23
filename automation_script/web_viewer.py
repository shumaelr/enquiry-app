"""
Web viewer for displaying processed enquiry documents.
Runs a simple HTTP server to show Excel outputs as HTML tables.
Supports file upload for PDF processing.
"""
import os
import json
import subprocess
import cgi
import tempfile
import shutil
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import parse_qs, urlparse, unquote, quote
import openpyxl
from datetime import datetime

# Configuration
WATCH_DIRECTORY = os.getenv("WATCH_DIRECTORY", "/Users/shumaelr/RealCode/Enquiry/2025.11.29R to Shumael - AI - BOQ, Sizing & SLD")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PORT = int(os.getenv("PORT", "8080"))


def get_all_excel_files():
    """Get all Excel files (templates and outputs) from the watched directories."""
    files = {"outputs": [], "templates": []}
    
    for folder in ["BOQ", "Sizing", "SLD"]:
        folder_path = os.path.join(WATCH_DIRECTORY, folder)
        if not os.path.exists(folder_path):
            continue
            
        for filename in os.listdir(folder_path):
            if filename.startswith("~$"):
                continue
                
            if filename.endswith(".xlsx"):
                filepath = os.path.join(folder_path, filename)
                stat = os.stat(filepath)
                modified_dt = datetime.fromtimestamp(stat.st_mtime)
                
                file_info = {
                    "folder": folder,
                    "filename": filename,
                    "filepath": filepath,
                    "modified": modified_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "modified_dt": modified_dt,
                    "date_folder": modified_dt.strftime("%Y-%m-%d"),
                    "time": modified_dt.strftime("%H:%M:%S"),
                    "size": f"{stat.st_size / 1024:.1f} KB"
                }
                
                if filename.startswith("Output_"):
                    # Parse source file from output filename
                    # Format: Output_<source_name>_YYYYMMDD_HHMMSS.xlsx
                    parts = filename[7:-5]  # Remove "Output_" and ".xlsx"
                    # Find the timestamp part (last 15 chars: YYYYMMDD_HHMMSS)
                    if len(parts) > 16 and parts[-15:-7].isdigit():
                        source_name = parts[:-16]  # Everything before _YYYYMMDD_HHMMSS
                        timestamp_str = parts[-15:]  # YYYYMMDD_HHMMSS
                    else:
                        source_name = parts
                        timestamp_str = ""
                    
                    file_info["source_name"] = source_name
                    file_info["timestamp_str"] = timestamp_str
                    files["outputs"].append(file_info)
                elif "Format" in filename:
                    files["templates"].append(file_info)
    
    # Sort by modified date, newest first
    files["outputs"].sort(key=lambda x: x["modified"], reverse=True)
    files["templates"].sort(key=lambda x: x["folder"])
    return files


def read_excel_fully(filepath):
    """Read entire Excel file and return as list of rows with all data."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        
        rows = []
        max_col = ws.max_column or 1
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row_data = []
            for cell in row:
                value = cell.value if cell.value is not None else ""
                row_data.append(str(value))
            rows.append(row_data)
        
        wb.close()
        return rows, max_col
        
    except Exception as e:
        return [[f"Error reading file: {e}"]], 1


def read_excel_to_html(filepath):
    """Convert an Excel file to HTML table with proper styling."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        
        # Determine if multi-column or vertical format
        is_multicolumn = False
        for row_idx in range(1, 30):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_c = ws.cell(row=row_idx, column=3).value
            if cell_a and cell_c:
                if "item" in str(cell_a).lower() or "question" in str(cell_a).lower():
                    is_multicolumn = True
                    break
        
        html = '<table class="excel-table">'
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(100, ws.max_row)), 1):
            # Determine if this is a header row
            is_header = row_idx == 1
            for cell in row:
                if cell.value and "item" in str(cell.value).lower():
                    is_header = True
                    break
            
            if is_header:
                html += '<tr class="header-row">'
            else:
                html += '<tr>'
            
            for col_idx, cell in enumerate(row, 1):
                value = cell.value if cell.value is not None else ""
                # Escape HTML
                value = str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                # Handle newlines
                value = value.replace("\n", "<br>")
                
                # Style based on position
                if is_header:
                    html += f'<th>{value}</th>'
                elif col_idx == 1 and value:
                    html += f'<td class="label-cell">{value}</td>'
                elif value and value != "None":
                    html += f'<td class="data-cell">{value}</td>'
                else:
                    html += f'<td class="empty-cell">{value}</td>'
            html += '</tr>'
        
        html += '</table>'
        wb.close()
        return html
        
    except Exception as e:
        return f'<div class="error">Error reading file: {e}</div>'


def generate_outputs_page():
    """Generate the main page showing all output files organized by different criteria."""
    files = get_all_excel_files()
    outputs = files["outputs"]
    templates = files["templates"]
    
    # Group outputs by different criteria
    by_folder = {"BOQ": [], "Sizing": [], "SLD": []}
    by_source = {}  # Group by source file name
    by_date = {}    # Group by date
    
    for out in outputs:
        by_folder[out["folder"]].append(out)
        
        # Group by source name
        source = out.get("source_name", "Unknown")
        if source not in by_source:
            by_source[source] = []
        by_source[source].append(out)
        
        # Group by date
        date = out.get("date_folder", "Unknown")
        if date not in by_date:
            by_date[date] = []
        by_date[date].append(out)
    
    # Build the sidebar with grouped views
    sidebar_html = ""
    
    # By Date section
    sidebar_html += '<div class="sidebar-section"><h3>📅 By Date</h3>'
    for date in sorted(by_date.keys(), reverse=True):
        date_outputs = by_date[date]
        count = len(date_outputs)
        sidebar_html += f'''<div class="sidebar-group" data-filter="date-{date}">
            <span class="group-label">📆 {date}</span>
            <span class="group-count">{count}</span>
        </div>'''
    sidebar_html += '</div>'
    
    # By Source section
    sidebar_html += '<div class="sidebar-section"><h3>📄 By Source File</h3>'
    for source in sorted(by_source.keys()):
        source_outputs = by_source[source]
        count = len(source_outputs)
        # Determine primary folder for badge
        folders = set(o["folder"] for o in source_outputs)
        folder = list(folders)[0] if len(folders) == 1 else "mixed"
        sidebar_html += f'''<div class="sidebar-group" data-filter="source-{source}">
            <span class="folder-badge {folder.lower()}">{folder}</span>
            <span class="group-label" title="{source}">{source[:25]}{'...' if len(source) > 25 else ''}</span>
            <span class="group-count">{count}</span>
        </div>'''
    sidebar_html += '</div>'
    
    # By Category section
    sidebar_html += '<div class="sidebar-section"><h3>📁 By Category</h3>'
    for folder in ["BOQ", "Sizing", "SLD"]:
        count = len(by_folder[folder])
        if count > 0:
            sidebar_html += f'''<div class="sidebar-group" data-filter="folder-{folder}">
                <span class="folder-badge {folder.lower()}">{folder}</span>
                <span class="group-label">{folder}</span>
                <span class="group-count">{count}</span>
            </div>'''
    sidebar_html += '</div>'
    
    # Templates section in sidebar
    if templates:
        sidebar_html += '<div class="sidebar-section"><h3>📋 Templates</h3>'
        for tmpl in templates:
            sidebar_html += f'''<a class="sidebar-group template-link" href="/view?file={quote(tmpl['filepath'], safe='')}">
                <span class="folder-badge {tmpl['folder'].lower()}">{tmpl['folder']}</span>
                <span class="group-label">{tmpl['filename'][:20]}</span>
            </a>'''
        sidebar_html += '</div>'
    
    # Build main content with all outputs as cards
    cards_html = ""
    if outputs:
        for out in outputs:
            source = out.get("source_name", "Unknown")
            date = out.get("date_folder", "Unknown")
            folder = out["folder"]
            time = out.get("time", "")
            
            # Read preview (first few rows)
            rows, max_col = read_excel_fully(out["filepath"])
            preview_rows = rows[:5] if len(rows) >= 5 else rows
            
            preview_html = '<table class="preview-table">'
            for row_idx, row in enumerate(preview_rows):
                preview_html += '<tr>'
                for col_idx, cell in enumerate(row[:4]):  # Max 4 columns in preview
                    value = cell[:50] + '...' if len(cell) > 50 else cell
                    value = value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    if row_idx == 0 or col_idx == 0:
                        preview_html += f'<td class="preview-header">{value}</td>'
                    else:
                        preview_html += f'<td>{value}</td>'
                preview_html += '</tr>'
            preview_html += '</table>'
            
            cards_html += f'''
            <div class="output-card" data-folder="{folder}" data-source="{source}" data-date="{date}">
                <div class="card-header">
                    <span class="folder-badge {folder.lower()}">{folder}</span>
                    <span class="card-time">🕐 {time}</span>
                </div>
                <div class="card-source">
                    <span class="source-icon">📄</span>
                    <span class="source-name" title="{source}">{source}</span>
                </div>
                <div class="card-preview">
                    {preview_html}
                </div>
                <div class="card-footer">
                    <span class="card-date">📅 {date}</span>
                    <span class="card-size">{out["size"]}</span>
                    <a href="/view?file={quote(out['filepath'], safe='')}" class="view-btn">View Full →</a>
                </div>
            </div>
            '''
    else:
        cards_html = '''
        <div class="empty-state">
            <div class="empty-icon">📭</div>
            <h3>No Output Files Yet</h3>
            <p>Upload and process some PDFs to see results here</p>
            <a href="/upload" class="upload-cta">📤 Upload PDFs Now</a>
        </div>
        '''
    
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Enquiry Document Processor</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #ffffff;
            min-height: 100vh;
            color: #2d3748;
        }}
        .layout {{
            display: flex;
            min-height: 100vh;
        }}
        
        /* Sidebar */
        .sidebar {{
            width: 280px;
            background: #f8fafa;
            border-right: 1px solid #d4eeef;
            padding: 20px;
            position: fixed;
            height: 100vh;
            overflow-y: auto;
        }}
        .sidebar-header {{
            background: linear-gradient(135deg, #7accc8 0%, #3b9daa 100%);
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            color: white;
        }}
        .sidebar-header h1 {{
            font-size: 18px;
            margin-bottom: 5px;
        }}
        .sidebar-header .subtitle {{
            font-size: 12px;
            opacity: 0.9;
        }}
        .sidebar-stats {{
            display: flex;
            gap: 10px;
            margin-top: 15px;
        }}
        .mini-stat {{
            flex: 1;
            background: rgba(255,255,255,0.2);
            padding: 8px;
            border-radius: 6px;
            text-align: center;
        }}
        .mini-stat-num {{
            font-size: 20px;
            font-weight: 700;
        }}
        .mini-stat-label {{
            font-size: 9px;
            text-transform: uppercase;
            opacity: 0.9;
        }}
        .sidebar-section {{
            margin-bottom: 25px;
        }}
        .sidebar-section h3 {{
            font-size: 11px;
            text-transform: uppercase;
            color: #718096;
            margin-bottom: 10px;
            letter-spacing: 0.5px;
        }}
        .sidebar-group {{
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 10px 12px;
            margin-bottom: 4px;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.2s;
            background: white;
            border: 1px solid transparent;
            text-decoration: none;
        }}
        .sidebar-group:hover {{
            background: #e8f6f7;
            border-color: #b8e0e3;
        }}
        .sidebar-group.active {{
            background: linear-gradient(135deg, #e8f6f7 0%, #d4eeef 100%);
            border-color: #7accc8;
        }}
        .template-link {{
            font-size: 12px;
        }}
        .group-label {{
            flex: 1;
            font-size: 13px;
            color: #2d7a87;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        .group-count {{
            background: #3b9daa;
            color: white;
            padding: 2px 8px;
            border-radius: 10px;
            font-size: 11px;
            font-weight: 600;
        }}
        .upload-btn {{
            display: block;
            width: 100%;
            padding: 12px;
            background: linear-gradient(135deg, #5eb5b7 0%, #3b9daa 100%);
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            text-align: center;
            text-decoration: none;
            margin-bottom: 15px;
            transition: all 0.2s;
        }}
        .upload-btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 10px rgba(59, 157, 170, 0.3);
        }}
        .show-all-btn {{
            display: block;
            width: 100%;
            padding: 10px;
            background: white;
            color: #3b9daa;
            border: 1px solid #b8e0e3;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 500;
            text-align: center;
            transition: all 0.2s;
        }}
        .show-all-btn:hover {{
            background: #f0f9fa;
            border-color: #7accc8;
        }}
        
        /* Main content */
        .main-content {{
            margin-left: 280px;
            flex: 1;
            padding: 20px;
        }}
        header {{ 
            background: linear-gradient(135deg, #7accc8 0%, #3b9daa 100%);
            padding: 15px 25px;
            border-radius: 12px;
            margin-bottom: 20px;
            box-shadow: 0 4px 15px rgba(59, 157, 170, 0.3);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        h1 {{ 
            font-size: 24px; 
            color: #ffffff;
        }}
        .header-actions {{
            display: flex;
            gap: 10px;
        }}
        .header-btn {{
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.4);
            padding: 10px 20px;
            border-radius: 8px;
            color: white;
            cursor: pointer;
            text-decoration: none;
            font-weight: 500;
            transition: all 0.2s;
        }}
        .header-btn:hover {{ background: rgba(255,255,255,0.3); }}
        
        .filter-info {{
            background: #e8f6f7;
            padding: 12px 16px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: none;
            align-items: center;
            justify-content: space-between;
        }}
        .filter-info.active {{
            display: flex;
        }}
        .filter-text {{
            color: #2d7a87;
            font-size: 14px;
        }}
        .clear-filter {{
            background: none;
            border: 1px solid #3b9daa;
            color: #3b9daa;
            padding: 5px 12px;
            border-radius: 5px;
            cursor: pointer;
            font-size: 12px;
        }}
        .clear-filter:hover {{
            background: #3b9daa;
            color: white;
        }}
        
        /* Cards grid */
        .cards-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(350px, 1fr));
            gap: 20px;
        }}
        .output-card {{
            background: white;
            border: 1px solid #d4eeef;
            border-radius: 12px;
            overflow: hidden;
            transition: all 0.2s;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }}
        .output-card:hover {{
            transform: translateY(-3px);
            box-shadow: 0 8px 20px rgba(59, 157, 170, 0.15);
            border-color: #7accc8;
        }}
        .output-card.hidden {{
            display: none;
        }}
        .card-header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 12px 15px;
            background: #f8fafa;
            border-bottom: 1px solid #e8f6f7;
        }}
        .card-time {{
            font-size: 12px;
            color: #718096;
        }}
        .card-source {{
            padding: 12px 15px;
            display: flex;
            align-items: center;
            gap: 8px;
            border-bottom: 1px solid #e8f6f7;
        }}
        .source-icon {{
            font-size: 20px;
        }}
        .source-name {{
            font-weight: 600;
            color: #2d7a87;
            font-size: 14px;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        .card-preview {{
            padding: 12px;
            background: #fafcfc;
            max-height: 150px;
            overflow: hidden;
        }}
        .preview-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 11px;
        }}
        .preview-table td {{
            padding: 4px 6px;
            border: 1px solid #e8f6f7;
            color: #4a5568;
            max-width: 100px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        .preview-table .preview-header {{
            background: #e8f6f7;
            font-weight: 600;
            color: #2d7a87;
        }}
        .card-footer {{
            display: flex;
            align-items: center;
            gap: 15px;
            padding: 12px 15px;
            background: white;
            border-top: 1px solid #e8f6f7;
        }}
        .card-date, .card-size {{
            font-size: 12px;
            color: #718096;
        }}
        .view-btn {{
            margin-left: auto;
            background: linear-gradient(135deg, #5eb5b7 0%, #3b9daa 100%);
            color: white;
            padding: 6px 14px;
            border-radius: 6px;
            text-decoration: none;
            font-size: 12px;
            font-weight: 500;
            transition: all 0.2s;
        }}
        .view-btn:hover {{
            transform: translateY(-1px);
            box-shadow: 0 3px 8px rgba(59, 157, 170, 0.3);
        }}
        
        .folder-badge {{
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 10px;
            font-weight: 600;
            text-transform: uppercase;
        }}
        .folder-badge.boq {{ background: #006994; color: #ffffff; }}
        .folder-badge.sizing {{ background: #3b9daa; color: #ffffff; }}
        .folder-badge.sld {{ background: #7accc8; color: #1a5f66; }}
        .folder-badge.mixed {{ background: #718096; color: #ffffff; }}
        
        .empty-state {{
            text-align: center;
            padding: 80px 40px;
            background: #f8fafa;
            border-radius: 12px;
            border: 2px dashed #d4eeef;
        }}
        .empty-icon {{
            font-size: 64px;
            margin-bottom: 20px;
        }}
        .empty-state h3 {{
            color: #2d7a87;
            margin-bottom: 10px;
        }}
        .empty-state p {{
            color: #718096;
            margin-bottom: 25px;
        }}
        .upload-cta {{
            display: inline-block;
            background: linear-gradient(135deg, #5eb5b7 0%, #3b9daa 100%);
            color: white;
            padding: 15px 30px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            transition: all 0.2s;
        }}
        .upload-cta:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 15px rgba(59, 157, 170, 0.4);
        }}
        
        .no-results {{
            text-align: center;
            padding: 60px;
            color: #718096;
        }}
        .no-results-icon {{
            font-size: 48px;
            margin-bottom: 15px;
        }}
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <div class="sidebar-header">
                <h1>📊 Enquiry Processor</h1>
                <div class="subtitle">AI-Powered Document Analysis</div>
                <div class="sidebar-stats">
                    <div class="mini-stat">
                        <div class="mini-stat-num">{len(outputs)}</div>
                        <div class="mini-stat-label">Outputs</div>
                    </div>
                    <div class="mini-stat">
                        <div class="mini-stat-num">{len(by_folder['BOQ'])}</div>
                        <div class="mini-stat-label">BOQ</div>
                    </div>
                    <div class="mini-stat">
                        <div class="mini-stat-num">{len(by_folder['Sizing'])}</div>
                        <div class="mini-stat-label">Sizing</div>
                    </div>
                    <div class="mini-stat">
                        <div class="mini-stat-num">{len(by_folder['SLD'])}</div>
                        <div class="mini-stat-label">SLD</div>
                    </div>
                </div>
            </div>
            
            <a href="/upload" class="upload-btn">📤 Upload & Process PDFs</a>
            <button class="show-all-btn" onclick="showAll()">📑 Show All Outputs</button>
            
            {sidebar_html}
        </aside>
        
        <main class="main-content">
            <header>
                <h1>📊 Processed Documents</h1>
                <div class="header-actions">
                    <a href="/upload" class="header-btn">📤 Upload More</a>
                    <button class="header-btn" onclick="location.reload()">🔄 Refresh</button>
                </div>
            </header>
            
            <div class="filter-info" id="filterInfo">
                <span class="filter-text" id="filterText">Showing all outputs</span>
                <button class="clear-filter" onclick="showAll()">Clear Filter</button>
            </div>
            
            <div class="cards-grid" id="cardsGrid">
                {cards_html}
            </div>
            
            <div class="no-results" id="noResults" style="display: none;">
                <div class="no-results-icon">🔍</div>
                <p>No outputs match this filter</p>
            </div>
        </main>
    </div>
    
    <script>
        // Filter functionality
        document.querySelectorAll('.sidebar-group:not(.template-link)').forEach(group => {{
            group.addEventListener('click', function() {{
                const filter = this.dataset.filter;
                if (!filter) return;
                const [type, ...valueParts] = filter.split('-');
                const value = valueParts.join('-');
                
                // Update active state
                document.querySelectorAll('.sidebar-group').forEach(g => g.classList.remove('active'));
                this.classList.add('active');
                
                // Filter cards
                let visibleCount = 0;
                document.querySelectorAll('.output-card').forEach(card => {{
                    let show = false;
                    if (type === 'folder') {{
                        show = card.dataset.folder === value;
                    }} else if (type === 'source') {{
                        show = card.dataset.source === value;
                    }} else if (type === 'date') {{
                        show = card.dataset.date === value;
                    }}
                    
                    if (show) {{
                        card.classList.remove('hidden');
                        visibleCount++;
                    }} else {{
                        card.classList.add('hidden');
                    }}
                }});
                
                // Update filter info
                document.getElementById('filterInfo').classList.add('active');
                let label = value;
                if (type === 'folder') label = value + ' category';
                else if (type === 'source') label = 'Source: ' + value;
                else if (type === 'date') label = 'Date: ' + value;
                document.getElementById('filterText').textContent = 'Showing ' + visibleCount + ' output(s) for ' + label;
                
                // Show no results message if needed
                document.getElementById('noResults').style.display = visibleCount === 0 ? 'block' : 'none';
            }});
        }});
        
        function showAll() {{
            document.querySelectorAll('.sidebar-group').forEach(g => g.classList.remove('active'));
            document.querySelectorAll('.output-card').forEach(card => card.classList.remove('hidden'));
            document.getElementById('filterInfo').classList.remove('active');
            document.getElementById('noResults').style.display = 'none';
        }}
    </script>
</body>
</html>'''


def generate_index_html():
    """Generate the main index page."""
    files = get_all_excel_files()
    
    # Output files section
    outputs_html = ""
    for out in files["outputs"]:
        outputs_html += f'''
        <tr class="clickable-row" data-href="/view?file={quote(out['filepath'], safe='')}">
            <td><span class="folder-badge {out["folder"].lower()}">{out["folder"]}</span></td>
            <td>{out["filename"]}</td>
            <td>{out["modified"]}</td>
            <td>{out["size"]}</td>
        </tr>
        '''
    
    if not outputs_html:
        outputs_html = '<tr><td colspan="4" class="no-data">No output files found. Process some PDFs first!</td></tr>'
    
    # Template files section
    templates_html = ""
    for tmpl in files["templates"]:
        templates_html += f'''
        <tr class="clickable-row" data-href="/view?file={quote(tmpl['filepath'], safe='')}">
            <td><span class="folder-badge {tmpl["folder"].lower()}">{tmpl["folder"]}</span></td>
            <td>{tmpl["filename"]}</td>
            <td>{tmpl["modified"]}</td>
            <td>{tmpl["size"]}</td>
        </tr>
        '''
    
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Enquiry Document Processor - Dashboard</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #ffffff;
            min-height: 100vh;
            color: #2d3748;
        }}
        .container {{ max-width: 1200px; margin: 0 auto; padding: 20px; }}
        header {{ 
            background: linear-gradient(135deg, #7accc8 0%, #3b9daa 100%);
            padding: 20px 30px;
            border-radius: 12px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(59, 157, 170, 0.3);
        }}
        h1 {{ 
            font-size: 28px; 
            color: #ffffff;
            margin-bottom: 5px;
        }}
        h2 {{
            font-size: 18px;
            color: #2d7a87;
            margin: 30px 0 15px 0;
            padding-bottom: 10px;
            border-bottom: 2px solid #d4eeef;
        }}
        .subtitle {{ color: rgba(255,255,255,0.9); font-size: 14px; }}
        .stats {{ 
            display: flex; 
            gap: 20px; 
            margin-top: 20px;
            flex-wrap: wrap;
        }}
        .stat-card {{
            background: rgba(255,255,255,0.15);
            padding: 15px 25px;
            border-radius: 10px;
            border: 1px solid rgba(255,255,255,0.2);
        }}
        .stat-number {{ font-size: 32px; font-weight: bold; color: #ffffff; }}
        .stat-label {{ font-size: 12px; color: rgba(255,255,255,0.8); text-transform: uppercase; }}
        .files-table {{
            width: 100%;
            border-collapse: collapse;
            background: #ffffff;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 4px 15px rgba(0,0,0,0.05);
            border: 1px solid #d4eeef;
        }}
        .files-table th {{
            text-align: left;
            padding: 15px 20px;
            background: linear-gradient(135deg, #e8f6f7 0%, #d4eeef 100%);
            font-weight: 600;
            color: #2d7a87;
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 1px;
        }}
        .files-table td {{
            padding: 15px 20px;
            border-bottom: 1px solid #e8f6f7;
            color: #4a5568;
        }}
        .clickable-row {{
            cursor: pointer;
            transition: background 0.2s;
        }}
        .clickable-row:hover {{
            background: #f0f9fa;
        }}
        .folder-badge {{
            padding: 5px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }}
        .folder-badge.boq {{ background: #006994; color: #ffffff; }}
        .folder-badge.sizing {{ background: #3b9daa; color: #ffffff; }}
        .folder-badge.sld {{ background: #7accc8; color: #1a5f66; }}
        .no-data {{ 
            text-align: center; 
            padding: 40px !important; 
            color: #a0aec0; 
        }}
        .btn-row {{
            display: flex;
            gap: 10px;
            margin-top: 20px;
        }}
        .action-btn {{
            background: linear-gradient(135deg, #5eb5b7 0%, #3b9daa 100%);
            border: none;
            padding: 12px 24px;
            border-radius: 8px;
            color: white;
            cursor: pointer;
            font-weight: 600;
            text-decoration: none;
            display: inline-block;
            box-shadow: 0 4px 10px rgba(59, 157, 170, 0.3);
            transition: all 0.2s;
        }}
        .action-btn:hover {{ 
            transform: translateY(-2px);
            box-shadow: 0 6px 15px rgba(59, 157, 170, 0.4);
        }}
        .action-btn.secondary {{
            background: #ffffff;
            border: 2px solid #b8e0e3;
            color: #3b9daa;
            box-shadow: none;
        }}
        .action-btn.secondary:hover {{
            background: #f0f9fa;
            border-color: #7accc8;
        }}
    </style>
    <script>
        document.addEventListener('DOMContentLoaded', function() {{
            document.querySelectorAll('.clickable-row').forEach(function(row) {{
                row.addEventListener('click', function() {{
                    window.location = this.dataset.href;
                }});
            }});
        }});
    </script>
</head>
<body>
    <div class="container">
        <header>
            <h1>📊 Enquiry Document Processor</h1>
            <p class="subtitle">Automated PDF analysis powered by Claude AI</p>
            <div class="stats">
                <div class="stat-card">
                    <div class="stat-number">{len(files["outputs"])}</div>
                    <div class="stat-label">Processed Files</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{len([o for o in files["outputs"] if o["folder"] == "BOQ"])}</div>
                    <div class="stat-label">BOQ Documents</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{len([o for o in files["outputs"] if o["folder"] == "Sizing"])}</div>
                    <div class="stat-label">Sizing Documents</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{len([o for o in files["outputs"] if o["folder"] == "SLD"])}</div>
                    <div class="stat-label">SLD Documents</div>
                </div>
            </div>
        </header>
        
        <div class="btn-row">
            <a href="/upload" class="action-btn">📤 Upload & Process PDF</a>
            <a href="/outputs" class="action-btn secondary">📑 View All Outputs</a>
            <button class="action-btn secondary" onclick="location.reload()">🔄 Refresh</button>
        </div>
        
        <h2>📄 Processed Output Files</h2>
        <table class="files-table">
            <thead>
                <tr>
                    <th>Category</th>
                    <th>File Name</th>
                    <th>Processed</th>
                    <th>Size</th>
                </tr>
            </thead>
            <tbody>
                {outputs_html}
            </tbody>
        </table>
        
        <h2>📋 Template Files</h2>
        <table class="files-table">
            <thead>
                <tr>
                    <th>Category</th>
                    <th>File Name</th>
                    <th>Modified</th>
                    <th>Size</th>
                </tr>
            </thead>
            <tbody>
                {templates_html}
            </tbody>
        </table>
    </div>
</body>
</html>'''


def generate_view_html(filepath):
    """Generate the file view page."""
    filename = os.path.basename(filepath)
    folder = os.path.basename(os.path.dirname(filepath))
    table_html = read_excel_to_html(filepath)
    
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{filename} - Enquiry Viewer</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #ffffff;
            min-height: 100vh;
            color: #2d3748;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        header {{ 
            background: linear-gradient(135deg, #7accc8 0%, #3b9daa 100%);
            padding: 20px 30px;
            border-radius: 12px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(59, 157, 170, 0.3);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .back-btn {{
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.4);
            padding: 10px 20px;
            border-radius: 8px;
            color: white;
            cursor: pointer;
            text-decoration: none;
            font-weight: 500;
            transition: all 0.2s;
        }}
        .back-btn:hover {{ background: rgba(255,255,255,0.3); }}
        h1 {{ font-size: 20px; color: #fff; }}
        .folder-badge {{
            padding: 5px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            margin-right: 10px;
        }}
        .folder-badge.boq {{ background: #006994; color: #ffffff; }}
        .folder-badge.sizing {{ background: #3b9daa; color: #ffffff; }}
        .folder-badge.sld {{ background: #7accc8; color: #1a5f66; }}
        .table-container {{
            overflow-x: auto;
            background: #ffffff;
            border-radius: 12px;
            padding: 20px;
            border: 1px solid #d4eeef;
            box-shadow: 0 4px 15px rgba(0,0,0,0.05);
        }}
        .excel-table {{
            width: 100%;
            border-collapse: collapse;
            min-width: 600px;
        }}
        .excel-table th {{
            padding: 12px 16px;
            background: linear-gradient(135deg, #5eb5b7 0%, #4a9ba8 100%);
            color: #ffffff;
            font-weight: 600;
            text-align: left;
            border-bottom: 2px solid #3b9daa;
            white-space: nowrap;
        }}
        .excel-table td {{
            padding: 12px 16px;
            border-bottom: 1px solid #e8f6f7;
            vertical-align: top;
        }}
        .excel-table tr:hover {{
            background: #f0f9fa;
        }}
        .header-row th {{
            background: linear-gradient(135deg, #3b9daa 0%, #2d7a87 100%);
            color: #fff;
        }}
        .label-cell {{
            font-weight: 600;
            color: #2d7a87;
            white-space: nowrap;
            min-width: 250px;
            max-width: 400px;
            background: #f0f9fa;
        }}
        .data-cell {{
            color: #1a5f66;
            max-width: 400px;
        }}
        .empty-cell {{
            color: #b8e0e3;
        }}
        .error {{
            background: #fff5f5;
            color: #c53030;
            padding: 20px;
            border-radius: 10px;
            border: 1px solid #fed7d7;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div>
                <span class="folder-badge {folder.lower()}">{folder}</span>
                <span style="font-size: 14px; color: rgba(255,255,255,0.9);">{filename}</span>
            </div>
            <a href="/" class="back-btn">← Back to Dashboard</a>
        </header>
        
        <div class="table-container">
            {table_html}
        </div>
    </div>
</body>
</html>'''


def generate_upload_html(message="", success=False, results=None):
    """Generate the file upload page."""
    message_html = ""
    if message:
        if success:
            message_html = f'<div class="message success">{message}</div>'
        else:
            message_html = f'<div class="message error">{message}</div>'
    
    # Generate results table if we have batch results
    results_html = ""
    if results:
        results_rows = ""
        for r in results:
            status_class = "success" if r.get('success') else "error"
            status_icon = "✅" if r.get('success') else "❌"
            results_rows += f'''<tr class="{status_class}">
                <td>{status_icon}</td>
                <td><span class="folder-badge {r.get('category', '').lower()}">{r.get('category', 'N/A')}</span></td>
                <td>{r.get('filename', 'Unknown')}</td>
                <td>{r.get('message', '')}</td>
            </tr>'''
        
        results_html = f'''<div class="results-container">
            <h3>📊 Processing Results</h3>
            <table class="results-table">
                <thead>
                    <tr>
                        <th style="width: 40px;">Status</th>
                        <th style="width: 80px;">Category</th>
                        <th>File</th>
                        <th>Result</th>
                    </tr>
                </thead>
                <tbody>
                    {results_rows}
                </tbody>
            </table>
        </div>'''
    
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Upload PDFs - Enquiry Document Processor</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #ffffff;
            min-height: 100vh;
            color: #2d3748;
        }}
        .container {{ max-width: 900px; margin: 0 auto; padding: 20px; }}
        header {{ 
            background: linear-gradient(135deg, #7accc8 0%, #3b9daa 100%);
            padding: 20px 30px;
            border-radius: 12px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(59, 157, 170, 0.3);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        h1 {{ font-size: 24px; color: #ffffff; }}
        h3 {{ color: #2d7a87; margin-bottom: 15px; }}
        .back-btn {{
            background: rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.4);
            padding: 10px 20px;
            border-radius: 8px;
            color: white;
            cursor: pointer;
            text-decoration: none;
            font-weight: 500;
            transition: all 0.2s;
        }}
        .back-btn:hover {{ background: rgba(255,255,255,0.3); }}
        
        .upload-modes {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 20px;
            margin-bottom: 30px;
        }}
        .mode-card {{
            background: #ffffff;
            border: 2px solid #d4eeef;
            border-radius: 12px;
            padding: 25px;
            cursor: pointer;
            transition: all 0.3s;
        }}
        .mode-card:hover {{
            border-color: #7accc8;
            background: #f0f9fa;
        }}
        .mode-card.selected {{
            border-color: #3b9daa;
            background: linear-gradient(135deg, #e8f6f7 0%, #d4eeef 100%);
        }}
        .mode-card .mode-icon {{
            font-size: 36px;
            margin-bottom: 10px;
        }}
        .mode-card .mode-title {{
            font-size: 18px;
            font-weight: 600;
            color: #2d7a87;
            margin-bottom: 5px;
        }}
        .mode-card .mode-desc {{
            font-size: 13px;
            color: #718096;
        }}
        
        .upload-section {{
            display: none;
        }}
        .upload-section.active {{
            display: block;
        }}
        
        .upload-container {{
            background: #ffffff;
            border: 2px dashed #b8e0e3;
            border-radius: 12px;
            padding: 40px;
            text-align: center;
            margin-bottom: 20px;
            transition: all 0.3s;
        }}
        .upload-container:hover {{
            border-color: #3b9daa;
            background: #f0f9fa;
        }}
        .upload-container.dragover {{
            border-color: #3b9daa;
            background: #e8f6f7;
        }}
        .upload-icon {{
            font-size: 48px;
            margin-bottom: 20px;
        }}
        .upload-text {{
            font-size: 18px;
            color: #4a5568;
            margin-bottom: 10px;
        }}
        .upload-subtext {{
            font-size: 14px;
            color: #a0aec0;
        }}
        
        .form-group {{
            margin-bottom: 20px;
            text-align: left;
        }}
        .form-group > label {{
            display: block;
            font-weight: 600;
            color: #2d7a87;
            margin-bottom: 8px;
        }}
        
        .submit-btn {{
            background: linear-gradient(135deg, #5eb5b7 0%, #3b9daa 100%);
            border: none;
            padding: 15px 40px;
            border-radius: 8px;
            color: white;
            cursor: pointer;
            font-weight: 600;
            font-size: 16px;
            box-shadow: 0 4px 10px rgba(59, 157, 170, 0.3);
            transition: all 0.2s;
            width: 100%;
        }}
        .submit-btn:hover {{ 
            transform: translateY(-2px);
            box-shadow: 0 6px 15px rgba(59, 157, 170, 0.4);
        }}
        .submit-btn:disabled {{
            opacity: 0.6;
            cursor: not-allowed;
            transform: none;
        }}
        
        .message {{
            padding: 15px 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            font-weight: 500;
        }}
        .message.success {{
            background: #e8f6f7;
            color: #1a5f66;
            border: 1px solid #7accc8;
        }}
        .message.error {{
            background: #fff5f5;
            color: #c53030;
            border: 1px solid #fed7d7;
        }}
        
        .processing {{
            display: none;
            text-align: center;
            padding: 30px;
        }}
        .processing.active {{
            display: block;
        }}
        .spinner {{
            width: 40px;
            height: 40px;
            border: 4px solid #d4eeef;
            border-top-color: #3b9daa;
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin: 0 auto 20px;
        }}
        @keyframes spin {{
            to {{ transform: rotate(360deg); }}
        }}
        
        .category-cards {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 15px;
            margin-bottom: 20px;
        }}
        .category-card {{
            padding: 20px;
            border: 2px solid #d4eeef;
            border-radius: 10px;
            cursor: pointer;
            text-align: center;
            transition: all 0.2s;
        }}
        .category-card:hover {{
            border-color: #7accc8;
            background: #f0f9fa;
        }}
        .category-card.selected {{
            border-color: #3b9daa;
            background: linear-gradient(135deg, #e8f6f7 0%, #d4eeef 100%);
        }}
        .category-card input {{
            display: none;
        }}
        .category-card .emoji {{
            font-size: 32px;
            margin-bottom: 10px;
        }}
        .category-card .name {{
            font-weight: 600;
            color: #2d7a87;
        }}
        .category-card .desc {{
            font-size: 12px;
            color: #718096;
            margin-top: 5px;
        }}
        
        .file-list {{
            background: #f8fafa;
            border: 1px solid #d4eeef;
            border-radius: 8px;
            padding: 15px;
            margin-top: 15px;
            max-height: 200px;
            overflow-y: auto;
        }}
        .file-item {{
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 8px 12px;
            background: white;
            border-radius: 6px;
            margin-bottom: 8px;
            font-size: 14px;
        }}
        .file-item:last-child {{ margin-bottom: 0; }}
        .file-item .file-path {{
            color: #718096;
            font-size: 12px;
        }}
        .file-item .file-name {{
            color: #2d7a87;
            font-weight: 500;
        }}
        .file-count {{
            font-weight: 600;
            color: #3b9daa;
            margin-bottom: 10px;
        }}
        
        .folder-badge {{
            display: inline-block;
            padding: 3px 10px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
        }}
        .folder-badge.boq {{ background: #e8f6f7; color: #2d7a87; }}
        .folder-badge.sizing {{ background: #d4eeef; color: #1a5f66; }}
        .folder-badge.sld {{ background: #b8e0e3; color: #2d7a87; }}
        
        .results-container {{
            background: #ffffff;
            border: 1px solid #d4eeef;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 30px;
        }}
        .results-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        .results-table th {{
            background: #f0f9fa;
            padding: 10px;
            text-align: left;
            font-size: 13px;
            color: #2d7a87;
            border-bottom: 2px solid #d4eeef;
        }}
        .results-table td {{
            padding: 10px;
            border-bottom: 1px solid #e8f6f7;
            font-size: 13px;
        }}
        .results-table tr.success td {{
            background: #f0faf0;
        }}
        .results-table tr.error td {{
            background: #fff8f8;
        }}
        
        .info-box {{
            background: #e8f6f7;
            border: 1px solid #7accc8;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 20px;
            font-size: 14px;
            color: #2d7a87;
        }}
        .info-box strong {{
            color: #1a5f66;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>📤 Upload & Process PDFs</h1>
            <a href="/" class="back-btn">← Back to Dashboard</a>
        </header>
        
        {message_html}
        {results_html}
        
        <div class="upload-modes">
            <div class="mode-card" id="modeMultiple" onclick="selectMode('multiple')">
                <div class="mode-icon">📑</div>
                <div class="mode-title">Multiple Files</div>
                <div class="mode-desc">Select multiple PDFs and choose a category for all of them</div>
            </div>
            <div class="mode-card" id="modeFolder" onclick="selectMode('folder')">
                <div class="mode-icon">📁</div>
                <div class="mode-title">Folder Upload</div>
                <div class="mode-desc">Upload a folder with BOQ/, Sizing/, SLD/ subfolders - categories auto-detected</div>
            </div>
        </div>
        
        <!-- Multiple Files Mode -->
        <div class="upload-section" id="sectionMultiple">
            <form method="POST" action="/upload" enctype="multipart/form-data" id="uploadFormMultiple">
                <input type="hidden" name="mode" value="multiple">
                
                <div class="form-group">
                    <label>Select Document Category:</label>
                    <div class="category-cards">
                        <label class="category-card" onclick="selectCategory(this, 'BOQ')">
                            <input type="radio" name="category" value="BOQ">
                            <div class="emoji">📋</div>
                            <div class="name">BOQ</div>
                            <div class="desc">Bill of Quantities</div>
                        </label>
                        <label class="category-card" onclick="selectCategory(this, 'Sizing')">
                            <input type="radio" name="category" value="Sizing">
                            <div class="emoji">📐</div>
                            <div class="name">Sizing</div>
                            <div class="desc">Solar System Sizing</div>
                        </label>
                        <label class="category-card" onclick="selectCategory(this, 'SLD')">
                            <input type="radio" name="category" value="SLD">
                            <div class="emoji">⚡</div>
                            <div class="name">SLD</div>
                            <div class="desc">Single Line Diagram</div>
                        </label>
                    </div>
                </div>
                
                <div class="upload-container" id="dropZoneMultiple">
                    <div class="upload-icon">📄</div>
                    <div class="upload-text">Drag & drop PDF files here</div>
                    <div class="upload-subtext">or click to browse - select multiple files</div>
                    <input type="file" name="pdf_files" id="fileInputMultiple" accept=".pdf" multiple style="display: none;">
                    <div id="fileListMultiple" class="file-list" style="display: none;"></div>
                </div>
                
                <button type="submit" class="submit-btn" id="submitBtnMultiple">
                    🚀 Process All Documents with AI
                </button>
            </form>
        </div>
        
        <!-- Folder Mode -->
        <div class="upload-section" id="sectionFolder">
            <form method="POST" action="/upload" enctype="multipart/form-data" id="uploadFormFolder">
                <input type="hidden" name="mode" value="folder">
                
                <div class="info-box">
                    <strong>📂 Folder Structure:</strong> Upload a folder containing <code>BOQ/</code>, <code>Sizing/</code>, and/or <code>SLD/</code> subfolders. 
                    Categories will be auto-detected from folder names.
                </div>
                
                <div class="upload-container" id="dropZoneFolder">
                    <div class="upload-icon">📁</div>
                    <div class="upload-text">Drag & drop a folder here</div>
                    <div class="upload-subtext">or click to browse folder</div>
                    <input type="file" name="folder_files" id="fileInputFolder" webkitdirectory directory multiple style="display: none;">
                    <div id="fileListFolder" class="file-list" style="display: none;"></div>
                </div>
                
                <button type="submit" class="submit-btn" id="submitBtnFolder">
                    🚀 Process All Documents with AI
                </button>
            </form>
        </div>
        
        <div class="processing" id="processingDiv">
            <div class="processing-layout">
                <div class="processing-left">
                    <div class="processing-header">
                        <div class="spinner"></div>
                        <div class="processing-title">Processing Documents</div>
                    </div>
                    <div class="progress-bar-container">
                        <div class="progress-bar" id="progressBar"></div>
                    </div>
                    <div class="progress-text" id="progressText">Preparing files...</div>
                    <div class="processing-steps">
                        <div class="step-item" id="step-upload"><span class="step-icon">📤</span> Uploading</div>
                        <div class="step-item" id="step-extract"><span class="step-icon">📄</span> Extracting PDF</div>
                        <div class="step-item" id="step-analyze"><span class="step-icon">🤖</span> AI Analysis</div>
                        <div class="step-item" id="step-generate"><span class="step-icon">📊</span> Generating Excel</div>
                    </div>
                </div>
                <div class="processing-right">
                    <div class="file-list-header">📁 Files Progress</div>
                    <div class="file-progress-list" id="fileProgressList"></div>
                </div>
            </div>
        </div>
    </div>
    
    <style>
        .processing {{
            display: none;
            background: white;
            border-radius: 16px;
            padding: 30px;
            margin-top: 30px;
            box-shadow: 0 4px 20px rgba(45, 122, 135, 0.15);
            border: 2px solid #7accc8;
        }}
        .processing.active {{
            display: block;
        }}
        .processing-layout {{
            display: flex;
            gap: 30px;
        }}
        .processing-left {{
            flex: 1;
            min-width: 300px;
        }}
        .processing-right {{
            flex: 1;
            min-width: 350px;
            border-left: 2px solid #e8f6f7;
            padding-left: 30px;
        }}
        .processing-header {{
            display: flex;
            align-items: center;
            gap: 15px;
            margin-bottom: 25px;
        }}
        .processing-title {{
            font-size: 22px;
            font-weight: 600;
            color: #2d7a87;
        }}
        .progress-bar-container {{
            background: #e8f6f7;
            border-radius: 10px;
            height: 24px;
            overflow: hidden;
            margin-bottom: 15px;
        }}
        .progress-bar {{
            height: 100%;
            background: linear-gradient(90deg, #7accc8, #3b9daa, #2d7a87);
            width: 0%;
            transition: width 0.3s ease;
            border-radius: 10px;
            position: relative;
        }}
        .progress-bar::after {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.3), transparent);
            animation: shimmer 1.5s infinite;
        }}
        @keyframes shimmer {{
            0% {{ transform: translateX(-100%); }}
            100% {{ transform: translateX(100%); }}
        }}
        .progress-text {{
            text-align: center;
            color: #2d7a87;
            font-size: 14px;
            margin-bottom: 30px;
            font-weight: 500;
        }}
        .processing-steps {{
            display: flex;
            flex-direction: column;
            gap: 12px;
        }}
        .step-item {{
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 12px 16px;
            background: #f8fafa;
            border-radius: 8px;
            font-size: 14px;
            color: #718096;
            transition: all 0.3s ease;
        }}
        .step-item.active {{
            background: #fff9e6;
            color: #92400e;
            font-weight: 500;
            border-left: 4px solid #f6ad55;
        }}
        .step-item.complete {{
            background: #e6fff2;
            color: #065f46;
            border-left: 4px solid #48bb78;
        }}
        .step-icon {{
            font-size: 18px;
        }}
        .file-list-header {{
            font-size: 16px;
            font-weight: 600;
            color: #2d7a87;
            margin-bottom: 15px;
            padding-bottom: 10px;
            border-bottom: 2px solid #e8f6f7;
        }}
        .file-progress-list {{
            max-height: 350px;
            overflow-y: auto;
        }}
        .file-progress-item {{
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 10px 12px;
            background: #f8fafa;
            border-radius: 8px;
            margin-bottom: 8px;
            transition: all 0.3s ease;
        }}
        .file-progress-item.processing {{
            background: #fff9e6;
            border-left: 4px solid #f6ad55;
        }}
        .file-progress-item.complete {{
            background: #e6fff2;
            border-left: 4px solid #48bb78;
        }}
        .file-progress-item.error {{
            background: #ffe6e6;
            border-left: 4px solid #f56565;
        }}
        .file-progress-item.waiting {{
            opacity: 0.5;
        }}
        .file-status-icon {{
            font-size: 20px;
            width: 30px;
            text-align: center;
        }}
        .file-info {{
            flex: 1;
        }}
        .file-name {{
            font-weight: 500;
            color: #2d3748;
            font-size: 14px;
        }}
        .file-step {{
            font-size: 12px;
            color: #718096;
            margin-top: 3px;
        }}
        .file-category {{
            font-size: 11px;
            padding: 3px 8px;
            border-radius: 4px;
            font-weight: 500;
        }}
        .file-category.boq {{ background: #fef3c7; color: #92400e; }}
        .file-category.sizing {{ background: #dbeafe; color: #1e40af; }}
        .file-category.sld {{ background: #d1fae5; color: #065f46; }}
        .file-category.unknown {{ background: #f3f4f6; color: #6b7280; }}
        
        @keyframes pulse {{
            0%, 100% {{ opacity: 1; }}
            50% {{ opacity: 0.5; }}
        }}
        .file-progress-item.processing .file-status-icon {{
            animation: pulse 1s infinite;
        }}
    </style>
    
    <script>
        let currentMode = null;
        let isProcessing = false;
        let allowNavigation = false;
        
        // Prevent navigation while processing
        window.addEventListener('beforeunload', function(e) {{
            if (isProcessing && !allowNavigation) {{
                e.preventDefault();
                e.returnValue = 'Files are still being processed. Are you sure you want to leave?';
                return e.returnValue;
            }}
        }});
        
        // Handle back button clicks
        document.addEventListener('click', function(e) {{
            if (isProcessing) {{
                const link = e.target.closest('a');
                if (link && link.href) {{
                    e.preventDefault();
                    e.stopPropagation();
                    if (confirm('Files are still being processed. Are you sure you want to leave? Processing will continue in the background.')) {{
                        allowNavigation = true;
                        isProcessing = false;
                        window.location.href = link.href;
                    }}
                    return false;
                }}
            }}
        }}, true);
        
        function selectMode(mode) {{
            currentMode = mode;
            document.querySelectorAll('.mode-card').forEach(c => c.classList.remove('selected'));
            document.querySelectorAll('.upload-section').forEach(s => s.classList.remove('active'));
            
            document.getElementById('mode' + mode.charAt(0).toUpperCase() + mode.slice(1)).classList.add('selected');
            document.getElementById('section' + mode.charAt(0).toUpperCase() + mode.slice(1)).classList.add('active');
        }}
        
        function selectCategory(card, category) {{
            document.querySelectorAll('.category-card').forEach(c => c.classList.remove('selected'));
            card.classList.add('selected');
            card.querySelector('input').checked = true;
        }}
        
        // Multiple files mode
        const dropZoneMultiple = document.getElementById('dropZoneMultiple');
        const fileInputMultiple = document.getElementById('fileInputMultiple');
        const fileListMultiple = document.getElementById('fileListMultiple');
        
        dropZoneMultiple.addEventListener('click', () => fileInputMultiple.click());
        
        dropZoneMultiple.addEventListener('dragover', (e) => {{
            e.preventDefault();
            dropZoneMultiple.classList.add('dragover');
        }});
        
        dropZoneMultiple.addEventListener('dragleave', () => {{
            dropZoneMultiple.classList.remove('dragover');
        }});
        
        dropZoneMultiple.addEventListener('drop', (e) => {{
            e.preventDefault();
            dropZoneMultiple.classList.remove('dragover');
            const files = Array.from(e.dataTransfer.files).filter(f => f.type === 'application/pdf');
            if (files.length > 0) {{
                const dt = new DataTransfer();
                files.forEach(f => dt.items.add(f));
                fileInputMultiple.files = dt.files;
                updateFileListMultiple(files);
            }}
        }});
        
        fileInputMultiple.addEventListener('change', () => {{
            const files = Array.from(fileInputMultiple.files);
            updateFileListMultiple(files);
        }});
        
        function updateFileListMultiple(files) {{
            if (files.length > 0) {{
                fileListMultiple.style.display = 'block';
                fileListMultiple.innerHTML = '<div class="file-count">📎 ' + files.length + ' file(s) selected:</div>' +
                    files.map(f => '<div class="file-item"><span class="file-name">📄 ' + f.name + '</span></div>').join('');
            }} else {{
                fileListMultiple.style.display = 'none';
            }}
        }}
        
        // Folder mode
        const dropZoneFolder = document.getElementById('dropZoneFolder');
        const fileInputFolder = document.getElementById('fileInputFolder');
        const fileListFolder = document.getElementById('fileListFolder');
        
        dropZoneFolder.addEventListener('click', () => fileInputFolder.click());
        
        dropZoneFolder.addEventListener('dragover', (e) => {{
            e.preventDefault();
            dropZoneFolder.classList.add('dragover');
        }});
        
        dropZoneFolder.addEventListener('dragleave', () => {{
            dropZoneFolder.classList.remove('dragover');
        }});
        
        dropZoneFolder.addEventListener('drop', async (e) => {{
            e.preventDefault();
            dropZoneFolder.classList.remove('dragover');
            
            const items = e.dataTransfer.items;
            const files = [];
            
            async function traverseDirectory(entry, path = '') {{
                if (entry.isFile) {{
                    return new Promise((resolve) => {{
                        entry.file((file) => {{
                            if (file.name.endsWith('.pdf')) {{
                                file.relativePath = path + file.name;
                                files.push(file);
                            }}
                            resolve();
                        }});
                    }});
                }} else if (entry.isDirectory) {{
                    const reader = entry.createReader();
                    return new Promise((resolve) => {{
                        reader.readEntries(async (entries) => {{
                            for (const e of entries) {{
                                await traverseDirectory(e, path + entry.name + '/');
                            }}
                            resolve();
                        }});
                    }});
                }}
            }}
            
            for (const item of items) {{
                const entry = item.webkitGetAsEntry();
                if (entry) {{
                    await traverseDirectory(entry);
                }}
            }}
            
            if (files.length > 0) {{
                updateFileListFolder(files);
            }}
        }});
        
        fileInputFolder.addEventListener('change', () => {{
            const files = Array.from(fileInputFolder.files).filter(f => f.name.endsWith('.pdf'));
            updateFileListFolder(files);
        }});
        
        function updateFileListFolder(files) {{
            if (files.length > 0) {{
                const byCategory = {{}};
                files.forEach(f => {{
                    const path = f.webkitRelativePath || f.relativePath || f.name;
                    let category = 'Unknown';
                    if (path.toLowerCase().includes('/boq/') || path.toLowerCase().startsWith('boq/')) category = 'BOQ';
                    else if (path.toLowerCase().includes('/sizing/') || path.toLowerCase().startsWith('sizing/')) category = 'Sizing';
                    else if (path.toLowerCase().includes('/sld/') || path.toLowerCase().startsWith('sld/')) category = 'SLD';
                    
                    if (!byCategory[category]) byCategory[category] = [];
                    byCategory[category].push({{name: f.name, path: path}});
                }});
                
                let html = '<div class="file-count">📎 ' + files.length + ' PDF(s) found:</div>';
                for (const [cat, catFiles] of Object.entries(byCategory)) {{
                    const badgeClass = cat.toLowerCase();
                    html += catFiles.map(f => 
                        '<div class="file-item"><span class="folder-badge ' + badgeClass + '">' + cat + '</span><span class="file-name">' + f.name + '</span><span class="file-path">' + f.path + '</span></div>'
                    ).join('');
                }}
                
                fileListFolder.style.display = 'block';
                fileListFolder.innerHTML = html;
            }} else {{
                fileListFolder.style.display = 'none';
            }}
        }}
        
        // Processing steps for animation
        const processingSteps = [
            {{ step: 'upload', text: 'Uploading file...', icon: '📤' }},
            {{ step: 'extract', text: 'Extracting text from PDF...', icon: '📄' }},
            {{ step: 'analyze', text: 'Analyzing with AI...', icon: '🤖' }},
            {{ step: 'generate', text: 'Generating Excel output...', icon: '📊' }},
            {{ step: 'complete', text: 'Complete!', icon: '✅' }}
        ];
        
        function resetStepIndicators() {{
            ['upload', 'extract', 'analyze', 'generate'].forEach(step => {{
                const el = document.getElementById('step-' + step);
                if (el) {{
                    el.className = 'step-item';
                }}
            }});
        }}
        
        function updateStepIndicator(currentStep) {{
            const steps = ['upload', 'extract', 'analyze', 'generate'];
            const currentIndex = steps.indexOf(currentStep);
            
            steps.forEach((step, index) => {{
                const el = document.getElementById('step-' + step);
                if (el) {{
                    if (index < currentIndex) {{
                        el.className = 'step-item complete';
                    }} else if (index === currentIndex) {{
                        el.className = 'step-item active';
                    }} else {{
                        el.className = 'step-item';
                    }}
                }}
            }});
        }}
        
        function initializeFileProgress(files, getCategory) {{
            const list = document.getElementById('fileProgressList');
            list.innerHTML = '';
            resetStepIndicators();
            
            files.forEach((file, index) => {{
                const category = getCategory(file);
                const categoryClass = category.toLowerCase();
                const item = document.createElement('div');
                item.className = 'file-progress-item waiting';
                item.id = 'file-progress-' + index;
                item.innerHTML = `
                    <div class="file-status-icon">⏳</div>
                    <div class="file-info">
                        <div class="file-name">${{file.name}}</div>
                        <div class="file-step">Waiting...</div>
                    </div>
                    <div class="file-category ${{categoryClass}}">${{category}}</div>
                `;
                list.appendChild(item);
            }});
        }}
        
        function updateFileProgress(index, step, isError = false) {{
            const item = document.getElementById('file-progress-' + index);
            if (!item) return;
            
            item.className = 'file-progress-item ' + (isError ? 'error' : (step === 'complete' ? 'complete' : 'processing'));
            
            const stepInfo = processingSteps.find(s => s.step === step) || {{ icon: '❌', text: 'Error' }};
            item.querySelector('.file-status-icon').textContent = isError ? '❌' : stepInfo.icon;
            item.querySelector('.file-step').textContent = isError ? step : stepInfo.text;
            
            // Update left side step indicators
            if (!isError && step !== 'complete') {{
                updateStepIndicator(step);
            }}
        }}
        
        function updateOverallProgress(current, total, currentFileName) {{
            const percent = Math.round((current / total) * 100);
            document.getElementById('progressBar').style.width = percent + '%';
            document.getElementById('progressText').textContent = `Processing file ${{current}} of ${{total}}: ${{currentFileName}}`;
        }}
        
        // Form submissions
        document.getElementById('uploadFormMultiple').addEventListener('submit', async function(e) {{
            e.preventDefault();
            
            const category = document.querySelector('input[name="category"]:checked');
            if (!category) {{
                alert('Please select a category');
                return;
            }}
            if (fileInputMultiple.files.length === 0) {{
                alert('Please select at least one PDF file');
                return;
            }}
            
            const files = Array.from(fileInputMultiple.files);
            const categoryValue = category.value;
            
            isProcessing = true;
            document.getElementById('submitBtnMultiple').disabled = true;
            document.getElementById('submitBtnMultiple').textContent = 'Processing...';
            document.getElementById('processingDiv').classList.add('active');
            
            initializeFileProgress(files, () => categoryValue);
            
            // Process files one by one for animation
            const formData = new FormData();
            formData.append('mode', 'multiple');
            formData.append('category', categoryValue);
            
            for (let i = 0; i < files.length; i++) {{
                formData.append('pdf_files', files[i]);
            }}
            
            // Simulate step-by-step progress while processing
            let currentFile = 0;
            const simulateProgress = setInterval(() => {{
                if (currentFile < files.length) {{
                    updateOverallProgress(currentFile + 1, files.length, files[currentFile].name);
                    updateFileProgress(currentFile, 'analyze');
                }}
            }}, 500);
            
            try {{
                const response = await fetch('/upload', {{
                    method: 'POST',
                    body: formData
                }});
                
                clearInterval(simulateProgress);
                
                // Mark all as complete before redirect
                for (let i = 0; i < files.length; i++) {{
                    updateFileProgress(i, 'complete');
                }}
                document.getElementById('progressBar').style.width = '100%';
                document.getElementById('progressText').textContent = 'All files processed!';
                
                await new Promise(resolve => setTimeout(resolve, 500));
                
                isProcessing = false;
                const html = await response.text();
                document.open();
                document.write(html);
                document.close();
            }} catch (error) {{
                clearInterval(simulateProgress);
                isProcessing = false;
                alert('Upload failed: ' + error.message);
                document.getElementById('submitBtnMultiple').disabled = false;
                document.getElementById('submitBtnMultiple').textContent = '🚀 Process All Documents with AI';
                document.getElementById('processingDiv').classList.remove('active');
            }}
        }});
        
        document.getElementById('uploadFormFolder').addEventListener('submit', async function(e) {{
            e.preventDefault();
            
            if (fileInputFolder.files.length === 0) {{
                alert('Please select a folder');
                return;
            }}
            
            // Get PDF files with their categories
            const allFiles = Array.from(fileInputFolder.files);
            const pdfFiles = allFiles.filter(f => f.name.toLowerCase().endsWith('.pdf'));
            
            if (pdfFiles.length === 0) {{
                alert('No PDF files found in the folder');
                return;
            }}
            
            isProcessing = true;
            document.getElementById('submitBtnFolder').disabled = true;
            document.getElementById('submitBtnFolder').textContent = 'Processing...';
            document.getElementById('processingDiv').classList.add('active');
            
            // Initialize file list with detected categories
            const getCategory = (file) => {{
                const path = (file.webkitRelativePath || file.name).toLowerCase();
                const parts = path.split('/');
                for (const part of parts) {{
                    if (part === 'boq') return 'BOQ';
                    if (part === 'sizing') return 'Sizing';
                    if (part === 'sld') return 'SLD';
                }}
                return 'Unknown';
            }};
            
            initializeFileProgress(pdfFiles, getCategory);
            
            // Create FormData
            const formData = new FormData();
            formData.append('mode', 'folder');
            
            for (const file of pdfFiles) {{
                const relativePath = file.webkitRelativePath || file.name;
                formData.append('folder_files', file, relativePath);
            }}
            
            // Simulate step-by-step progress
            let stepIndex = 0;
            const steps = ['upload', 'extract', 'analyze', 'generate'];
            let currentFileIndex = 0;
            
            updateOverallProgress(1, pdfFiles.length, pdfFiles[0].name);
            updateFileProgress(0, 'upload');
            
            const simulateProgress = setInterval(() => {{
                stepIndex++;
                if (stepIndex < steps.length) {{
                    updateFileProgress(currentFileIndex, steps[stepIndex]);
                }} else {{
                    // Move to next file
                    updateFileProgress(currentFileIndex, 'complete');
                    currentFileIndex++;
                    stepIndex = 0;
                    if (currentFileIndex < pdfFiles.length) {{
                        updateOverallProgress(currentFileIndex + 1, pdfFiles.length, pdfFiles[currentFileIndex].name);
                        updateFileProgress(currentFileIndex, steps[0]);
                    }}
                }}
            }}, 800);
            
            try {{
                const response = await fetch('/upload', {{
                    method: 'POST',
                    body: formData
                }});
                
                clearInterval(simulateProgress);
                
                // Mark all as complete
                for (let i = 0; i < pdfFiles.length; i++) {{
                    updateFileProgress(i, 'complete');
                }}
                document.getElementById('progressBar').style.width = '100%';
                document.getElementById('progressText').textContent = 'All files processed successfully!';
                
                await new Promise(resolve => setTimeout(resolve, 800));
                
                isProcessing = false;
                const html = await response.text();
                document.open();
                document.write(html);
                document.close();
            }} catch (error) {{
                clearInterval(simulateProgress);
                isProcessing = false;
                alert('Upload failed: ' + error.message);
                document.getElementById('submitBtnFolder').disabled = false;
                document.getElementById('submitBtnFolder').textContent = '🚀 Process All Documents with AI';
                document.getElementById('processingDiv').classList.remove('active');
            }}
        }});
    </script>
</body>
</html>'''


def process_uploaded_pdf(category, pdf_path):
    """Process an uploaded PDF using the process_single.py script."""
    try:
        # Build the relative path expected by process_single.py
        rel_path = f"{category}/{os.path.basename(pdf_path)}"
        
        # Run the processing script
        result = subprocess.run(
            ["python3", "process_single.py", rel_path],
            cwd=SCRIPT_DIR,
            capture_output=True,
            text=True,
            timeout=120
        )
        
        if result.returncode == 0:
            # Extract output filename from the result
            output = result.stdout
            if "Output saved to:" in output:
                output_file = output.split("Output saved to:")[-1].strip()
                return True, f"Successfully processed! Output: {output_file}"
            return True, "Processing complete!"
        else:
            return False, f"Processing failed: {result.stderr or result.stdout}"
            
    except subprocess.TimeoutExpired:
        return False, "Processing timed out after 120 seconds"
    except Exception as e:
        return False, f"Error during processing: {str(e)}"


class RequestHandler(SimpleHTTPRequestHandler):
    """Custom HTTP request handler."""
    
    def do_GET(self):
        parsed = urlparse(self.path)
        
        if parsed.path == "/" or parsed.path == "" or parsed.path == "/outputs":
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(generate_outputs_page().encode())
        
        elif parsed.path == "/upload":
            self.send_response(200)
            self.send_header("Content-type", "text/html")
            self.end_headers()
            self.wfile.write(generate_upload_html().encode())
            
        elif parsed.path == "/view":
            query = parse_qs(parsed.query)
            filepath = query.get("file", [""])[0]
            # URL decode the filepath
            filepath = unquote(filepath)
            
            if filepath and os.path.exists(filepath):
                self.send_response(200)
                self.send_header("Content-type", "text/html")
                self.end_headers()
                self.wfile.write(generate_view_html(filepath).encode())
            else:
                self.send_response(404)
                self.send_header("Content-type", "text/html")
                self.end_headers()
                self.wfile.write(f"<h1>File not found</h1><p>{filepath}</p>".encode())
                
        else:
            self.send_error(404, "Not found")
    
    def do_POST(self):
        parsed = urlparse(self.path)
        
        if parsed.path == "/upload":
            # Parse multipart form data
            content_type = self.headers.get('Content-Type')
            if not content_type or 'multipart/form-data' not in content_type:
                self.send_response(400)
                self.send_header("Content-type", "text/html")
                self.end_headers()
                self.wfile.write(generate_upload_html("Invalid form submission", False).encode())
                return
            
            # Parse the multipart data
            try:
                # Get content length
                content_length = int(self.headers.get('Content-Length', 0))
                
                # Read the body
                body = self.rfile.read(content_length)
                
                # Parse boundary from content type
                boundary = content_type.split('boundary=')[1].encode()
                
                # Parse the multipart data manually
                parts = body.split(b'--' + boundary)
                
                mode = None
                category = None
                files_data = []  # List of (filename, relative_path, data, detected_category)
                
                for part in parts:
                    if b'name="mode"' in part:
                        lines = part.split(b'\r\n')
                        for i, line in enumerate(lines):
                            if line == b'' and i + 1 < len(lines):
                                mode = lines[i + 1].decode().strip()
                                break
                    
                    elif b'name="category"' in part:
                        lines = part.split(b'\r\n')
                        for i, line in enumerate(lines):
                            if line == b'' and i + 1 < len(lines):
                                category = lines[i + 1].decode().strip()
                                break
                    
                    elif b'name="pdf_files"' in part or b'name="folder_files"' in part or b'name="pdf_file"' in part:
                        # Extract filename and relative path
                        header_end = part.find(b'\r\n\r\n')
                        if header_end != -1:
                            header = part[:header_end].decode('utf-8', errors='replace')
                            filename = None
                            if 'filename="' in header:
                                filename = header.split('filename="')[1].split('"')[0]
                            
                            if filename and filename.lower().endswith('.pdf'):
                                pdf_data = part[header_end + 4:]
                                # Remove trailing boundary markers
                                if pdf_data.endswith(b'\r\n'):
                                    pdf_data = pdf_data[:-2]
                                if pdf_data.endswith(b'--'):
                                    pdf_data = pdf_data[:-2]
                                if pdf_data.endswith(b'\r\n'):
                                    pdf_data = pdf_data[:-2]
                                
                                # Skip empty files
                                if len(pdf_data) < 100:
                                    continue
                                
                                # Detect category from path for folder uploads
                                # Normalize path separators (Windows uses \, Unix uses /)
                                detected_cat = None
                                path_normalized = filename.replace('\\', '/').lower()
                                
                                # Check for category folders anywhere in the path
                                # This handles paths like "FolderName/BOQ/file.pdf" or "BOQ/file.pdf"
                                path_parts = path_normalized.split('/')
                                for part_name in path_parts:
                                    if part_name == 'boq':
                                        detected_cat = 'BOQ'
                                        break
                                    elif part_name == 'sizing':
                                        detected_cat = 'Sizing'
                                        break
                                    elif part_name == 'sld':
                                        detected_cat = 'SLD'
                                        break
                                
                                # Get just the filename without path
                                base_filename = os.path.basename(filename.replace('\\', '/'))
                                
                                print(f"[DEBUG] File: {filename} -> Category: {detected_cat}, Base: {base_filename}")
                                
                                files_data.append({
                                    'filename': base_filename,
                                    'relative_path': filename,
                                    'data': pdf_data,
                                    'detected_category': detected_cat
                                })
                
                # Determine processing mode
                if not mode:
                    mode = 'multiple' if category else 'folder'
                
                if len(files_data) == 0:
                    self.send_response(200)
                    self.send_header("Content-type", "text/html")
                    self.end_headers()
                    self.wfile.write(generate_upload_html("No PDF files found in upload", False).encode())
                    return
                
                # Process all files
                results = []
                success_count = 0
                
                for file_info in files_data:
                    # Determine category
                    file_category = file_info['detected_category'] or category
                    
                    if not file_category:
                        results.append({
                            'filename': file_info['filename'],
                            'category': 'Unknown',
                            'success': False,
                            'message': 'Could not determine category - place in BOQ/, Sizing/, or SLD/ folder'
                        })
                        continue
                    
                    # Save the file
                    dest_folder = os.path.join(WATCH_DIRECTORY, file_category)
                    dest_path = os.path.join(dest_folder, file_info['filename'])
                    
                    try:
                        with open(dest_path, 'wb') as f:
                            f.write(file_info['data'])
                        
                        print(f"Saved: {dest_path}")
                        
                        # Process the PDF
                        success, message = process_uploaded_pdf(file_category, dest_path)
                        
                        results.append({
                            'filename': file_info['filename'],
                            'category': file_category,
                            'success': success,
                            'message': message
                        })
                        
                        if success:
                            success_count += 1
                            
                    except Exception as e:
                        results.append({
                            'filename': file_info['filename'],
                            'category': file_category,
                            'success': False,
                            'message': f'Error: {str(e)}'
                        })
                
                # Generate response
                total = len(results)
                if success_count == total:
                    summary_msg = f"✅ Successfully processed all {total} file(s)!"
                    overall_success = True
                elif success_count > 0:
                    summary_msg = f"⚠️ Processed {success_count} of {total} file(s) successfully"
                    overall_success = True
                else:
                    summary_msg = f"❌ Failed to process all {total} file(s)"
                    overall_success = False
                
                self.send_response(200)
                self.send_header("Content-type", "text/html")
                self.end_headers()
                self.wfile.write(generate_upload_html(summary_msg, overall_success, results).encode())
                
            except Exception as e:
                print(f"Upload error: {e}")
                import traceback
                traceback.print_exc()
                self.send_response(200)
                self.send_header("Content-type", "text/html")
                self.end_headers()
                self.wfile.write(generate_upload_html(f"Upload error: {str(e)}", False).encode())
        else:
            self.send_error(404, "Not found")
    
    def log_message(self, format, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]}")


def main():
    """Start the web server."""
    print("=" * 60)
    print("ENQUIRY DOCUMENT VIEWER")
    print("=" * 60)
    print(f"Starting server at http://localhost:{PORT}")
    print("Press Ctrl+C to stop")
    print("=" * 60)
    
    server = HTTPServer(("", PORT), RequestHandler)
    
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        server.shutdown()


if __name__ == "__main__":
    main()
